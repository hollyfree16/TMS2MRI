"""
Parse NextStim export text (.nbe/.txt) into:
  - landmarks.csv
  - targets.csv
  - targets.xlsx  (one tab per Sequence ID + "AVG EF Max" summary tab)
  - metadata.txt
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


ID_RE = re.compile(r"^\d+(?:\.\d+)+\.$")
NUM_RE = re.compile(r"^-?\d+(?:\.\d+)?$")
SEQUENCE_ID_RE = re.compile(r"^(?:Defined series )?Sequence ID:\s*(.+)$")


TARGET_FIELDS = [
    "id", "time_ms", "stim_type", "inter_pulse_int_ms",
    "first_intens_pct", "second_intens_pct", "target_id", "rep_stim_id",
    "peeling_depth_mm", "user_resp",
    "coil_loc_x", "coil_loc_y", "coil_loc_z",
    "coil_norm_x", "coil_norm_y", "coil_norm_z",
    "coil_dir_x", "coil_dir_y", "coil_dir_z",
    "ef_max_loc_x", "ef_max_loc_y", "ef_max_loc_z",
    "ef_max_value_vm", "ef_at_target_vm",
]

TARGET_HEADERS = [
    "ID", "Time (ms)", "Stim. Type", "Inter Pulse Int. (ms)",
    "First Intens. (%%)", "Second Intens. (%%)", "Target ID", "Rep. Stim. ID",
    "Peeling Depth (mm)", "User Resp.",
    "Coil Loc. X", "Coil Loc. Y", "Coil Loc. Z",
    "Coil Normal X", "Coil Normal Y", "Coil Normal Z",
    "Coil Dir. X", "Coil Dir. Y", "Coil Dir. Z",
    "EF Max Loc. X", "EF Max Loc. Y", "EF Max Loc. Z",
    "EF Max Value (V/m)", "EF at Target (V/m)",
]

EF_MAX_VALUE_FIELD = "ef_max_value_vm"


def split_cells(line: str) -> list[str]:
    return [c.strip() for c in line.rstrip("\n").split("\t")]

def non_empty_cells(cells: Iterable[str]) -> list[str]:
    return [c for c in cells if c != ""]

def is_number(text: str) -> bool:
    return bool(NUM_RE.match(text))

def get_sequence_prefix(target_id: str) -> str:
    """'1.2.3.' -> '1.2'"""
    parts = target_id.rstrip(".").split(".")
    return ".".join(parts[:2]) if len(parts) >= 2 else parts[0]


def parse_landmarks(lines):
    landmarks, consumed = [], set()
    coordinate_system = None

    for line in lines:
        s = line.strip()
        if s.startswith("Coordinate system:"):
            coordinate_system = s.split(":", 1)[1].strip()

    start = None
    for idx, line in enumerate(lines):
        if line.strip() == "Landmarks (mm)":
            start = idx + 1
            break
    if start is None:
        return landmarks, consumed, coordinate_system

    for idx in range(start, len(lines)):
        if lines[idx].strip() == "Stimulation targets (mm)":
            break
        cells = non_empty_cells(split_cells(lines[idx]))
        if len(cells) >= 4 and all(is_number(cells[i]) for i in range(3)):
            landmarks.append({
                "coordinate_system": coordinate_system or "",
                "x": cells[0], "y": cells[1], "z": cells[2],
                "landmark_type": cells[3],
            })
            consumed.add(idx)

    return landmarks, consumed, coordinate_system


def parse_targets(lines):
    targets, consumed = [], set()
    for idx, line in enumerate(lines):
        cells = non_empty_cells(split_cells(line))
        if len(cells) < 3:
            continue
        if not ID_RE.match(cells[0]):
            continue
        if not is_number(cells[1]):
            continue
        if cells[2] not in {"Single", "Paired", "Burst", "Train"}:
            continue
        values = cells[:len(TARGET_FIELDS)]
        if len(values) < len(TARGET_FIELDS):
            values.extend([""] * (len(TARGET_FIELDS) - len(values)))
        targets.append(dict(zip(TARGET_FIELDS, values)))
        consumed.add(idx)
    return targets, consumed


def parse_sequence_ids(lines: list[str]) -> list[str]:
    """Return ordered, deduplicated list of sequence IDs (trailing dot stripped)."""
    seen, ordered = set(), []
    for line in lines:
        m = SEQUENCE_ID_RE.match(line.strip())
        if m:
            raw = m.group(1).strip().rstrip(".")
            if raw not in seen:
                seen.add(raw)
                ordered.append(raw)
    return ordered


def group_targets_by_sequence(targets, sequence_ids):
    groups = {seq: [] for seq in sequence_ids}
    for row in targets:
        prefix = get_sequence_prefix(row["id"])
        if prefix not in groups:
            groups[prefix] = []
        groups[prefix].append(row)
    return groups


def safe_float(value: str):
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def write_targets_xlsx(path: Path, targets, sequence_ids: list[str]) -> None:
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed — skipping targets.xlsx. "
              "Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    def apply_header(ws, ncols):
        for c in range(1, ncols + 1):
            ws.cell(row=1, column=c).font = bold

    def autofit(ws):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max((len(str(cell.value)) for cell in col if cell.value), default=8)
            ws.column_dimensions[letter].width = min(width + 4, 30)

    groups = group_targets_by_sequence(targets, sequence_ids)
    ordered = [s for s in sequence_ids if groups.get(s)]
    for k in groups:
        if k not in ordered and groups[k]:
            ordered.append(k)

    # ── AVG EF Max summary sheet (first sheet) ───────────────────────────────
    avg_ws = wb.active
    avg_ws.title = "AVG EF Max"
    avg_ws.append(["Sequence ID", "AVG EF Max Value (V/m)", "N Stimulations"])
    apply_header(avg_ws, 3)

    for seq_id in ordered:
        rows   = groups[seq_id]
        values = [v for v in (safe_float(r[EF_MAX_VALUE_FIELD]) for r in rows) if v is not None]
        avg    = round(sum(values) / len(values), 4) if values else None
        avg_ws.append([seq_id, avg if avg is not None else "N/A", len(rows)])

    autofit(avg_ws)

    # ── One sheet per sequence ────────────────────────────────────────────────
    for seq_id in ordered:
        ws = wb.create_sheet(title=seq_id[:31])
        ws.append(TARGET_HEADERS)
        apply_header(ws, len(TARGET_HEADERS))
        for row in groups[seq_id]:
            ws.append([row.get(f, "") for f in TARGET_FIELDS])
        autofit(ws)

    wb.save(path)


def write_csv(path, fieldnames, rows):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_metadata_file(path, lines, excluded):
    kept = [l for i, l in enumerate(lines) if i not in excluded]
    path.write_text("\n".join(kept) + "\n", encoding="utf-8")


def output_filename(base, ext, subject_id):
    return f"{base}_{subject_id}.{ext}" if subject_id else f"{base}.{ext}"


def parse_nbe(input_path: Path, output_dir: Path, subject_id=None) -> None:
    if subject_id:
        output_dir = output_dir / subject_id / "xnbe"
    output_dir.mkdir(parents=True, exist_ok=True)
    raw   = input_path.read_text(encoding="ascii", errors="replace")
    lines = raw.splitlines()

    landmarks, landmark_lines, _ = parse_landmarks(lines)
    targets,   target_lines      = parse_targets(lines)
    sequence_ids                 = parse_sequence_ids(lines)

    lm_path   = output_dir / output_filename("landmarks", "csv",  subject_id)
    tgt_path  = output_dir / output_filename("targets",   "csv",  subject_id)
    xlsx_path = output_dir / output_filename("targets",   "xlsx", subject_id)
    meta_path = output_dir / output_filename("metadata",  "txt",  subject_id)

    write_csv(lm_path,  ["coordinate_system", "x", "y", "z", "landmark_type"], landmarks)
    write_csv(tgt_path, TARGET_FIELDS, targets)
    write_targets_xlsx(xlsx_path, targets, sequence_ids)
    write_metadata_file(meta_path, lines, landmark_lines | target_lines)

    print(f"Written: {lm_path}   ({len(landmarks)} rows)")
    print(f"Written: {tgt_path}  ({len(targets)} rows)")
    if OPENPYXL_AVAILABLE:
        print(f"Written: {xlsx_path} ({len(sequence_ids)} sequence tab(s) + AVG EF Max)")
    print(f"Written: {meta_path}")


def parse_args():
    p = argparse.ArgumentParser(
        description="Parse NextStim export (.nbe/.txt) into landmarks, targets, and metadata.")
    p.add_argument("-i", "--input",      required=True, help="Input .nbe or .txt file.")
    p.add_argument("-o", "--output-dir", help="Output directory (default: <input_dir>/nbe_output).")
    p.add_argument("-s", "--subject-id", help="Optional subject ID appended to output filenames.")
    return p.parse_args()


if __name__ == "__main__":
    args       = parse_args()
    input_path = Path(args.input)
    output_dir = Path(args.output_dir) if args.output_dir else input_path.parent / "nbe_output"

    if not input_path.exists():
        print(f"Error: file not found: {input_path}")
        sys.exit(1)

    parse_nbe(input_path, output_dir, args.subject_id)