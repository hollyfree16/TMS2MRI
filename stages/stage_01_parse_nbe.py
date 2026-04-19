"""
stages/01_parse_nbe.py
======================
Stage 01: Parse a NextStim .nbe export into:
  - landmarks_<subject_id>.csv
  - targets_<subject_id>.csv
  - targets_<subject_id>.xlsx  (one tab per Sequence ID + AVG EF Max summary)

Adapted from the original parse_nbe.py — logic unchanged, IO routed
through PathManifest and all print() replaced with the stage logger.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from utils.io import PathManifest
from utils.logger import get_stage_logger

log = get_stage_logger("parse_nbe")

# ---------------------------------------------------------------------------
# Regexes / constants
# ---------------------------------------------------------------------------

ID_RE          = re.compile(r"^\d+(?:\.\d+)+\.$")
NUM_RE         = re.compile(r"^-?\d+(?:\.\d+)?$")
SEQUENCE_ID_RE = re.compile(r"^(?:Defined series )?Sequence ID:\s*(.+)$")

TARGET_FIELDS = [
    "id", "time_ms", "stim_type", "inter_pulse_int_ms",
    "first_intens_pct", "second_intens_pct", "target_id", "rep_stim_id",
    "peeling_depth_mm", "user_resp",
    "coil_loc_x", "coil_loc_y", "coil_loc_z",
    "coil_norm_x", "coil_norm_y", "coil_norm_z",
    "coil_dir_x",  "coil_dir_y",  "coil_dir_z",
    "ef_max_loc_x", "ef_max_loc_y", "ef_max_loc_z",
    "ef_max_value_vm", "ef_at_target_vm",
]

TARGET_HEADERS = [
    "ID", "Time (ms)", "Stim. Type", "Inter Pulse Int. (ms)",
    "First Intens. (%%)", "Second Intens. (%%)", "Target ID", "Rep. Stim. ID",
    "Peeling Depth (mm)", "User Resp.",
    "Coil Loc. X", "Coil Loc. Y", "Coil Loc. Z",
    "Coil Normal X", "Coil Normal Y", "Coil Normal Z",
    "Coil Dir. X",  "Coil Dir. Y",  "Coil Dir. Z",
    "EF Max Loc. X", "EF Max Loc. Y", "EF Max Loc. Z",
    "EF Max Value (V/m)", "EF at Target (V/m)",
]

EF_MAX_VALUE_FIELD = "ef_max_value_vm"


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _split_cells(line: str) -> list[str]:
    return [c.strip() for c in line.rstrip("\n").split("\t")]

def _non_empty(cells: Iterable[str]) -> list[str]:
    return [c for c in cells if c != ""]

def _is_number(text: str) -> bool:
    return bool(NUM_RE.match(text))

def _seq_prefix(target_id: str) -> str:
    parts = target_id.rstrip(".").split(".")
    return ".".join(parts[:2]) if len(parts) >= 2 else parts[0]

def _safe_float(value: str):
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _parse_landmarks(lines: list[str]):
    landmarks, consumed = [], set()
    coordinate_system = None

    for line in lines:
        s = line.strip()
        if s.startswith("Coordinate system:"):
            coordinate_system = s.split(":", 1)[1].strip()

    start = None
    for idx, line in enumerate(lines):
        if line.strip() == "Landmarks (mm)":
            start = idx + 1
            break
    if start is None:
        return landmarks, consumed, coordinate_system

    for idx in range(start, len(lines)):
        if lines[idx].strip() == "Stimulation targets (mm)":
            break
        cells = _non_empty(_split_cells(lines[idx]))
        if len(cells) >= 4 and all(_is_number(cells[i]) for i in range(3)):
            landmarks.append({
                "coordinate_system": coordinate_system or "",
                "x": cells[0], "y": cells[1], "z": cells[2],
                "landmark_type": cells[3],
            })
            consumed.add(idx)

    return landmarks, consumed, coordinate_system


def _parse_targets(lines: list[str]):
    targets, consumed = [], set()
    for idx, line in enumerate(lines):
        cells = _non_empty(_split_cells(line))
        if len(cells) < 3:
            continue
        if not ID_RE.match(cells[0]):
            continue
        if not _is_number(cells[1]):
            continue
        if cells[2] not in {"Single", "Paired", "Burst", "Train"}:
            continue
        values = cells[:len(TARGET_FIELDS)]
        if len(values) < len(TARGET_FIELDS):
            values.extend([""] * (len(TARGET_FIELDS) - len(values)))
        targets.append(dict(zip(TARGET_FIELDS, values)))
        consumed.add(idx)
    return targets, consumed


def _parse_sequence_ids(lines: list[str]) -> list[str]:
    seen, ordered = set(), []
    for line in lines:
        m = SEQUENCE_ID_RE.match(line.strip())
        if m:
            raw = m.group(1).strip().rstrip(".")
            if raw not in seen:
                seen.add(raw)
                ordered.append(raw)
    return ordered


def _group_targets(targets, sequence_ids):
    groups = {seq: [] for seq in sequence_ids}
    for row in targets:
        prefix = _seq_prefix(row["id"])
        groups.setdefault(prefix, []).append(row)
    return groups


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def _write_xlsx(path: Path, targets, sequence_ids: list[str]) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping targets.xlsx. "
                    "Install with: pip install openpyxl")
        return

    wb   = openpyxl.Workbook()
    bold = Font(bold=True)

    def _header(ws, ncols):
        for c in range(1, ncols + 1):
            ws.cell(row=1, column=c).font = bold

    def _autofit(ws):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width  = max((len(str(cell.value)) for cell in col if cell.value), default=8)
            ws.column_dimensions[letter].width = min(width + 4, 30)

    groups  = _group_targets(targets, sequence_ids)
    ordered = [s for s in sequence_ids if groups.get(s)]
    for k, v in groups.items():
        if k not in ordered and v:
            ordered.append(k)

    # AVG EF Max summary sheet
    avg_ws = wb.active
    avg_ws.title = "AVG EF Max"
    avg_ws.append(["Sequence ID", "AVG EF Max Value (V/m)", "N Stimulations"])
    _header(avg_ws, 3)

    for seq_id in ordered:
        rows   = groups[seq_id]
        values = [v for v in (_safe_float(r[EF_MAX_VALUE_FIELD]) for r in rows) if v is not None]
        avg    = round(sum(values) / len(values), 4) if values else None
        avg_ws.append([seq_id, avg if avg is not None else "N/A", len(rows)])
    _autofit(avg_ws)

    # One sheet per sequence
    for seq_id in ordered:
        ws = wb.create_sheet(title=seq_id[:31])
        ws.append(TARGET_HEADERS)
        _header(ws, len(TARGET_HEADERS))
        for row in groups[seq_id]:
            ws.append([row.get(f, "") for f in TARGET_FIELDS])
        _autofit(ws)

    wb.save(path)


# ---------------------------------------------------------------------------
# Stage entry point
# ---------------------------------------------------------------------------

def run(args, paths: PathManifest) -> None:
    """
    Parse args.nbe and write landmarks CSV, targets CSV, and targets XLSX
    to paths.xnbe_dir.
    """
    log.info("Parsing: %s", paths.nbe)

    raw   = paths.nbe.read_text(encoding="ascii", errors="replace")
    lines = raw.splitlines()

    landmarks, lm_consumed,   _ = _parse_landmarks(lines)
    targets,   tgt_consumed     = _parse_targets(lines)
    sequence_ids                = _parse_sequence_ids(lines)

    log.info("%d landmarks, %d stimulation events, %d sequence IDs",
             len(landmarks), len(targets), len(sequence_ids))

    paths.xnbe_dir.mkdir(parents=True, exist_ok=True)

    # Landmarks CSV
    with paths.landmarks_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["coordinate_system", "x", "y", "z", "landmark_type"])
        w.writeheader()
        w.writerows(landmarks)
    log.info("Landmarks CSV: %s", paths.landmarks_csv)

    # Targets CSV
    with paths.targets_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=TARGET_FIELDS)
        w.writeheader()
        w.writerows(targets)
    log.info("Targets  CSV : %s", paths.targets_csv)

    # Targets XLSX
    _write_xlsx(paths.targets_xlsx, targets, sequence_ids)
    if paths.targets_xlsx.exists():
        log.info("Targets XLSX : %s  (%d sequence tab(s))",
                 paths.targets_xlsx, len(sequence_ids))

    # Metadata (remaining lines not consumed by landmark/target parsers)
    consumed = lm_consumed | tgt_consumed
    meta_path = paths.xnbe_dir / f"metadata_{args.subject_id}.txt"
    kept = [l for i, l in enumerate(lines) if i not in consumed]
    meta_path.write_text("\n".join(kept) + "\n", encoding="utf-8")
    log.debug("Metadata     : %s", meta_path)
